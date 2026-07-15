from pathlib import Path

from openpyxl import Workbook

from rag_app.documents import DocumentProcessor


def _create_sales_workbook(path: Path) -> None:
    workbook = Workbook()

    dashboard = workbook.active
    dashboard.title = "Dashboard"
    dashboard["A1"] = "Коммерческий дашборд — 2026"
    dashboard["A2"] = "Сводные показатели и помесячная выручка"
    dashboard["A4"] = "Общая выручка"
    dashboard["B4"] = 85_735_400
    dashboard["A5"] = "Крупнейшая операция"
    dashboard["B5"] = 3_391_500
    dashboard["A7"] = "Месяц"
    dashboard["B7"] = "Выручка"
    dashboard.append(["Январь 2026", 12_336_000])
    dashboard.append(["Март 2026", 17_925_500])
    for row in (12, 13):
        for column in range(1, 5):
            dashboard.cell(row=row, column=column, value="Контрольный текст после таблицы")

    summary = workbook.create_sheet("Monthly Summary")
    summary.append(["Месяц", "Выручка", "Количество операций"])
    summary.append(["Январь 2026", 12_336_000, 10])
    summary.append(["Февраль 2026", 14_681_300, 10])
    summary.append(["Март 2026", 17_925_500, 10])

    reference = workbook.create_sheet("Reference")
    reference.append(["Поле", "Описание"])
    reference.append(["Маркер", "SALES-QUARTZ-3472"])
    reference.append(["Уникальная максимальная выручка", "3 391 500 ₽"])

    workbook.save(path)


def test_excel_preserves_independent_dashboard_blocks(tmp_path: Path) -> None:
    path = tmp_path / "sales.xlsx"
    _create_sales_workbook(path)

    documents = DocumentProcessor(chunk_size=300, chunk_overlap=0).load(str(path))

    dashboard_summary = next(
        document
        for document in documents
        if document.metadata["sheet"] == "Dashboard"
        and document.metadata["block_type"] == "key_value"
    )
    assert "Общая выручка: 85735400" in dashboard_summary.page_content
    assert "Крупнейшая операция: 3391500" in dashboard_summary.page_content
    control_chunks = [
        item for item in documents if "Контрольный текст после таблицы" in item.page_content
    ]
    assert len(control_chunks) == 1
    assert control_chunks[0].metadata["block_type"] == "text"
    assert len(documents) == len({item.page_content for item in documents})


def test_excel_builds_aggregate_chunks_with_row_context(tmp_path: Path) -> None:
    path = tmp_path / "sales.xlsx"
    _create_sales_workbook(path)

    documents = DocumentProcessor(chunk_size=300, chunk_overlap=0).load(str(path))

    revenue_aggregate = next(
        document
        for document in documents
        if document.metadata["sheet"] == "Monthly Summary"
        and document.metadata["block_type"] == "aggregate"
        and document.metadata["aggregate_column"] == "Выручка"
    )
    assert "максимальная месячная величина" in revenue_aggregate.page_content
    assert 'Максимум по показателю "Выручка": 17 925 500' in revenue_aggregate.page_content
    assert "Месяц: Март 2026" in revenue_aggregate.page_content
    assert 'Сумма по показателю "Выручка": 44 942 800' in revenue_aggregate.page_content


def test_excel_converts_field_description_table_to_facts(tmp_path: Path) -> None:
    path = tmp_path / "sales.xlsx"
    _create_sales_workbook(path)

    documents = DocumentProcessor(chunk_size=300, chunk_overlap=0).load(str(path))

    reference = next(
        document
        for document in documents
        if document.metadata["sheet"] == "Reference"
        and document.metadata["block_type"] == "key_value"
    )
    assert "Маркер: SALES-QUARTZ-3472" in reference.page_content
    assert "Уникальная максимальная выручка: 3 391 500 ₽" in reference.page_content
    assert "Поле:" not in reference.page_content
    assert all("_prechunked" not in document.metadata for document in documents)
